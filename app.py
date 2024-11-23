import os
from flask import Flask, request, render_template, send_file, jsonify
import cv2
import numpy as np
import pytesseract
import pandas as pd
from werkzeug.utils import secure_filename
import logging
import re
from datetime import datetime
from PIL import Image, ImageEnhance
import math
import sys
import openpyxl
import json

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('app.log')
    ]
)
logger = logging.getLogger(__name__)

# Set Tesseract path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def enhance_image(image_path):
    """Enhanced image preprocessing for better OCR accuracy."""
    # Read image with OpenCV
    img = cv2.imread(image_path)
    
    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Apply multiple preprocessing techniques
    
    # 1. Advanced Noise Reduction
    denoised = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
    
    # 2. Adaptive Thresholding with different parameters
    binary = cv2.adaptiveThreshold(
        denoised,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        15,  # Increased block size
        8    # Increased constant
    )
    
    # 3. Morphological operations to clean up text
    kernel = np.ones((1, 1), np.uint8)
    cleaned = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
    
    # 4. Scale image (make it larger)
    scale_percent = 200
    width = int(cleaned.shape[1] * scale_percent / 100)
    height = int(cleaned.shape[0] * scale_percent / 100)
    scaled = cv2.resize(cleaned, (width, height), interpolation=cv2.INTER_CUBIC)
    
    return scaled

def detect_table_structure(image):
    """Detect table structure in the image."""
    # Detect horizontal and vertical lines
    horizontal = np.copy(image)
    vertical = np.copy(image)
    
    # Specify size on horizontal axis
    cols = horizontal.shape[1]
    horizontal_size = cols // 30
    horizontalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (horizontal_size, 1))
    horizontal = cv2.erode(horizontal, horizontalStructure)
    horizontal = cv2.dilate(horizontal, horizontalStructure)
    
    # Specify size on vertical axis
    rows = vertical.shape[0]
    vertical_size = rows // 30
    verticalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (1, vertical_size))
    vertical = cv2.erode(vertical, verticalStructure)
    vertical = cv2.dilate(vertical, verticalStructure)
    
    # Create a mask of the table structure
    table_mask = cv2.bitwise_or(horizontal, vertical)
    
    # Find contours in the mask
    contours, _ = cv2.findContours(table_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # Sort contours by area
    contours = sorted(contours, key=cv2.contourArea, reverse=True)
    
    return contours

def extract_text_from_region(image, region):
    """Extract text from a specific region of the image."""
    x, y, w, h = region
    roi = image[y:y+h, x:x+w]
    
    # Apply additional preprocessing to the region
    roi = cv2.GaussianBlur(roi, (3, 3), 0)
    roi = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    
    # Use advanced Tesseract configuration
    custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz,.-/() '
    text = pytesseract.image_to_string(roi, config=custom_config)
    
    return text.strip()

def extract_account_info(text):
    """Extract account information from the text using improved patterns."""
    account_info = {
        'bank_name': '',
        'account_name': '',
        'account_number': ''
    }
    
    # Split text into lines and remove empty lines
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    # Bank name patterns
    bank_patterns = [
        r'(?i)(HDFC\s*BANK|ICICI\s*BANK|SBI|STATE\s*BANK\s*OF\s*INDIA|AXIS\s*BANK|KOTAK\s*MAHINDRA\s*BANK|YES\s*BANK|PUNJAB\s*NATIONAL\s*BANK|PNB|BANK\s*OF\s*BARODA|BOB)',
    ]
    
    # Account number patterns
    account_number_patterns = [
        r'(?i)A/C\s*(?:No\.?|Number|#)?\s*:?\s*(\d[\d\s*-]*\d)',
        r'(?i)Account\s*(?:No\.?|Number|#)?\s*:?\s*(\d[\d\s*-]*\d)',
        r'(?i)(?:No\.?|Number|#)?\s*:?\s*(\d{8,})',  # At least 8 digits
        r'\b(\d{8,})\b'  # Standalone number with at least 8 digits
    ]
    
    # Account name patterns
    account_name_patterns = [
        r'(?i)Account\s*Name\s*:?\s*([A-Za-z\s\.]+)',
        r'(?i)Name\s*:?\s*([A-Za-z\s\.]+)',
        r'(?i)Customer\s*Name\s*:?\s*([A-Za-z\s\.]+)',
        r'(?i)(?:Mr\.|Mrs\.|Ms\.|Dr\.)\s*([A-Za-z\s\.]+)'
    ]
    
    # Process each line
    for line in lines:
        # Bank Name
        if not account_info['bank_name']:
            for pattern in bank_patterns:
                match = re.search(pattern, line)
                if match:
                    account_info['bank_name'] = match.group(1).strip()
                    break
        
        # Account Number
        if not account_info['account_number']:
            for pattern in account_number_patterns:
                match = re.search(pattern, line)
                if match:
                    # Clean up the account number
                    acc_num = match.group(1).replace(' ', '').replace('-', '')
                    if len(acc_num) >= 8:  # Minimum length validation
                        account_info['account_number'] = acc_num
                        break
        
        # Account Name
        if not account_info['account_name']:
            for pattern in account_name_patterns:
                match = re.search(pattern, line)
                if match:
                    name = match.group(1).strip()
                    # Validate name (at least 2 words, each at least 2 chars)
                    words = [w for w in name.split() if len(w) >= 2]
                    if len(words) >= 2:
                        account_info['account_name'] = ' '.join(words)
                        break
    
    return account_info

def extract_transactions(text):
    """Extract transaction details with improved patterns."""
    transactions = []
    
    # Split text into lines
    lines = text.split('\n')
    
    # Transaction patterns
    date_pattern = r'(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})'
    amount_pattern = r'(?:Rs\.?|INR)?\s*([\d,]+\.?\d*)'
    
    # Process each line
    current_date = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Try to find a date
        date_match = re.search(date_pattern, line)
        if date_match:
            current_date = date_match.group(1)
            
            # Look for amounts in the same line
            amounts = re.findall(amount_pattern, line)
            description_parts = re.split(amount_pattern, line)[0]
            description = re.sub(date_pattern, '', description_parts).strip()
            
            if amounts:
                # Determine which amounts are withdrawals, deposits, and balance
                if len(amounts) >= 3:
                    transaction = {
                        'date': current_date,
                        'description': description,
                        'withdraw': amounts[0] if float(amounts[0].replace(',', '')) > 0 else '',
                        'deposit': amounts[1] if len(amounts) > 1 and float(amounts[1].replace(',', '')) > 0 else '',
                        'balance': amounts[-1]
                    }
                    transactions.append(transaction)
                elif len(amounts) == 2:
                    # Assume it's either withdraw/balance or deposit/balance
                    amount_val = float(amounts[0].replace(',', ''))
                    transaction = {
                        'date': current_date,
                        'description': description,
                        'withdraw': amounts[0] if amount_val > 0 else '',
                        'deposit': '',
                        'balance': amounts[1]
                    }
                    transactions.append(transaction)
    
    return transactions

def extract_data_from_image(image_path):
    """Main function to extract data from the bank statement image."""
    # Enhance image
    enhanced_image = enhance_image(image_path)
    
    # Use multiple OCR configurations for better accuracy
    configs = [
        '--oem 3 --psm 6',  # Assume uniform text block
        '--oem 3 --psm 4',  # Assume single column
        '--oem 3 --psm 3',  # Fully automatic
    ]
    
    all_text = ""
    for config in configs:
        text = pytesseract.image_to_string(enhanced_image, config=config)
        all_text += "\n" + text
    
    # Extract information
    account_info = extract_account_info(all_text)
    transactions = extract_transactions(all_text)
    
    return {
        'account_info': account_info,
        'transactions': transactions
    }

def process_transaction_row(row_data, transactions):
    """Process a row of transaction data."""
    if len(row_data) >= 5:  # Ensure we have all required fields
        try:
            # Try to parse date
            date_pattern = r'\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}'
            date_match = re.search(date_pattern, row_data[0])
            if date_match:
                date = date_match.group()
                description = ' '.join(row_data[1:-3])
                amounts = row_data[-3:]
                
                transaction = {
                    'date': date,
                    'description': description,
                    'withdraw': amounts[0] if is_number(amounts[0]) else '',
                    'deposit': amounts[1] if is_number(amounts[1]) else '',
                    'balance': amounts[2] if is_number(amounts[2]) else ''
                }
                transactions.append(transaction)
        except Exception as e:
            logger.error(f"Error processing transaction row: {e}")

def is_number(s):
    """Check if string can be converted to number."""
    try:
        float(s.replace(',', ''))
        return True
    except:
        return False

def extract_date(text):
    """Extract statement date from the text."""
    date_patterns = [
        r'(?i)statement\s+(?:date|period|for).{0,20}?(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})',
        r'(?i)as\s+of.{0,20}?(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})',
        r'(?i)date.{0,20}?(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})',
        r'(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})'
    ]
    
    text_block = ' '.join(text.split('\n'))
    
    for pattern in date_patterns:
        match = re.search(pattern, text_block)
        if match:
            try:
                date_str = match.group(1)
                # Normalize separators
                date_str = re.sub(r'[./]', '-', date_str)
                
                # Try different date formats
                for fmt in ['%d-%m-%Y', '%d-%m-%y']:
                    try:
                        return datetime.strptime(date_str, fmt).strftime('%d-%m-%Y')
                    except ValueError:
                        continue
            except Exception as e:
                logger.error(f"Error parsing date: {e}")
    
    return "Date not found"

def create_excel(data, output_path):
    """Create a formatted Excel file from the extracted data."""
    try:
        # Create a new workbook and select the active sheet
        workbook = openpyxl.Workbook()
        
        # Create Account Info sheet
        account_sheet = workbook.active
        account_sheet.title = "Account Info"
        
        # Style for headers
        header_style = openpyxl.styles.NamedStyle(name='header_style')
        header_style.font = openpyxl.styles.Font(bold=True, size=12)
        header_style.fill = openpyxl.styles.PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        header_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # Add account information
        account_headers = ['Field', 'Value']
        account_sheet.append(account_headers)
        for cell in account_sheet[1]:
            cell.style = header_style
        
        account_info = data.get('account_info', {})
        rows = [
            ['Bank Name', account_info.get('bank_name', '')],
            ['Account Name', account_info.get('account_name', '')],
            ['Account Number', account_info.get('account_number', '')],
            ['Statement Date', data.get('issue_date', '')]
        ]
        
        for row in rows:
            account_sheet.append(row)
        
        # Adjust column widths
        for column in account_sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            account_sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Create Transactions sheet
        trans_sheet = workbook.create_sheet(title='Transactions')
        
        # Add transaction headers
        trans_headers = ['Date', 'Description', 'Withdrawal', 'Deposit', 'Balance']
        trans_sheet.append(trans_headers)
        for cell in trans_sheet[1]:
            cell.style = header_style
        
        # Add transactions
        transactions = data.get('transactions', [])
        for trans in transactions:
            row = [
                trans.get('date', ''),
                trans.get('description', ''),
                trans.get('withdraw', ''),
                trans.get('deposit', ''),
                trans.get('balance', '')
            ]
            trans_sheet.append(row)
        
        # Format currency columns
        currency_format = openpyxl.styles.numbers.BUILTIN_FORMATS[44]  # Accounting format
        for col_letter in ['C', 'D', 'E']:  # Withdrawal, Deposit, Balance columns
            for cell in trans_sheet[col_letter]:
                if cell.row > 1:  # Skip header
                    try:
                        cell.value = float(str(cell.value).replace(',', ''))
                        cell.number_format = currency_format
                    except (ValueError, TypeError):
                        pass
        
        # Adjust transaction column widths
        for column in trans_sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            trans_sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(output_path)
        return True
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        return False

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health_check():
    """Health check endpoint to verify system status."""
    status = {
        'status': 'healthy',
        'tesseract': verify_tesseract(),
        'opencv': cv2.__version__,
        'upload_dir': os.path.exists(app.config['UPLOAD_FOLDER'])
    }
    return jsonify(status)

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp')):
            return jsonify({'error': 'Invalid file type'}), 400
        
        # Store original filename without extension for later use
        original_filename = os.path.splitext(secure_filename(file.filename))[0]
        
        # Save the file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Process the image
        enhanced_image = enhance_image(filepath)
        
        # Extract text using OCR with improved configuration
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(enhanced_image, config=custom_config)
        
        # Extract data
        data = extract_data_from_image(filepath)
        data['issue_date'] = extract_date(text)
        
        # Store the extracted data along with original filename
        extracted_data = {
            'data': data,
            'original_filename': original_filename
        }
        
        # Save as session data
        temp_data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_data.json')
        with open(temp_data_path, 'w') as f:
            json.dump(extracted_data, f)
        
        return jsonify(data)
        
    except Exception as e:
        logger.error(f"Error processing file: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        # Clean up the uploaded image
        try:
            if 'filepath' in locals():
                os.remove(filepath)
        except Exception as e:
            logger.error(f"Error cleaning up file: {e}")

@app.route('/convert', methods=['POST'])
def convert_to_excel():
    try:
        # Read the temporary data
        temp_data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_data.json')
        with open(temp_data_path, 'r') as f:
            extracted_data = json.load(f)
        
        data = extracted_data['data']
        original_filename = extracted_data['original_filename']
        
        # Create Excel file with original filename
        excel_filename = f"{original_filename}.xlsx"
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        
        if create_excel(data, excel_path):
            return send_file(
                excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=excel_filename
            )
        else:
            return jsonify({'error': 'Failed to create Excel file'}), 500
            
    except Exception as e:
        logger.error(f"Error converting to Excel: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        # Clean up temporary files
        try:
            if 'temp_data_path' in locals():
                os.remove(temp_data_path)
            if 'excel_path' in locals():
                os.remove(excel_path)
        except Exception as e:
            logger.error(f"Error cleaning up files: {e}")

def verify_tesseract():
    """Verify Tesseract installation and version."""
    try:
        version = pytesseract.get_tesseract_version()
        logger.info(f"Tesseract version: {version}")
        return True
    except Exception as e:
        logger.error(f"Tesseract not properly installed: {e}")
        return False

if __name__ == '__main__':
    # Verify system requirements
    if not verify_tesseract():
        logger.error("Tesseract OCR is not properly installed. Please verify installation.")
        sys.exit(1)
        
    logger.info("Starting Bank Statement OCR Converter...")
    logger.info(f"Upload directory: {os.path.abspath(app.config['UPLOAD_FOLDER'])}")
    
    app.run(debug=True)
